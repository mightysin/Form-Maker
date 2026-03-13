import io
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
from copy import copy
import streamlit as st

def copy_worksheet_format_and_data(source_ws, target_ws):
    """將範本的儲存格內容、寬高、合併格式、顏色與字體完美複製到新的工作表中"""
    for mcr in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(mcr))
        
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
        
    for row_idx, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = row_dim.height
            
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
                
    for img in source_ws._images:
        new_img = copy(img)
        target_ws.add_image(new_img, img.anchor)

def generate_excel(client_name, export_date, subtotal, tax, grand_total, cart, selected_notes, uploaded_file=None):
    template_wb = openpyxl.load_workbook("blank_form.xlsx")
    source_ws = template_wb.active
    
    if uploaded_file is not None:
        wb = openpyxl.load_workbook(uploaded_file)
        
        sheet_title = f"{export_date.month:02d}{export_date.day:02d}_報價"
        base_title = sheet_title
        counter = 1
        while sheet_title in wb.sheetnames:
            sheet_title = f"{base_title}_{counter}"
            counter += 1
            
        ws = wb.create_sheet(title=sheet_title)
        copy_worksheet_format_and_data(source_ws, ws)
        wb.active = ws  
    else:
        wb = template_wb
        ws = wb.active
        if client_name:
            ws.title = str(client_name)[:31] 

    ws.cell(row=6, column=2).value = client_name
    
    minguo_year = export_date.year - 1911
    date_str = f"{minguo_year}/{export_date.month}/{export_date.day}"
    ws.cell(row=6, column=6).value = date_str
    
    current_row = 8
    for idx, item in enumerate(cart):
        if "小計" in str(ws.cell(row=current_row, column=5).value).strip():
            st.warning("項目數量超過表單預留空間，部分項目可能未匯出。")
            break
            
        ws.cell(row=current_row, column=1).value = idx + 1               
        ws.cell(row=current_row, column=2).value = item.get("品名", "")  
        ws.cell(row=current_row, column=3).value = item.get("數量", 0)   
        ws.cell(row=current_row, column=4).value = item.get("單位", "")  
        ws.cell(row=current_row, column=5).value = item.get("單價", 0)   
        ws.cell(row=current_row, column=6).value = item.get("金額", 0)   
        current_row += 1

    subtotal_row = -1
    for r in range(current_row, current_row + 100):
        cell_val = str(ws.cell(row=r, column=5).value).strip()
        
        if "小計" in cell_val:
            subtotal_row = r
            ws.cell(row=r, column=6).value = subtotal       
            ws.cell(row=r+1, column=6).value = tax          
            ws.cell(row=r+2, column=6).value = grand_total  
            break

    if subtotal_row != -1 and subtotal_row > current_row:
        ranges_to_unmerge = []
        for m_range in ws.merged_cells.ranges:
            if not (m_range.max_row < current_row or m_range.min_row >= subtotal_row):
                ranges_to_unmerge.append(m_range)
        for m_range in ranges_to_unmerge:
            ws.unmerge_cells(str(m_range))
            
        rows_to_delete = subtotal_row - current_row
        ws.delete_rows(current_row, amount=rows_to_delete)
        
        new_subtotal_row = current_row
        
        right_align = Alignment(horizontal='right')
        for i in range(3):
            ws.cell(row=new_subtotal_row + i, column=5).alignment = right_align
        
        notes_row = -1
        for r in range(new_subtotal_row + 3, new_subtotal_row + 15):
            if str(ws.cell(row=r, column=1).value).strip() != "":
                notes_row = r
                break
        
        if notes_row != -1:
            notes_text = ""
            for i, note in enumerate(selected_notes):
                notes_text += f"{i+1}. {note}\n"
            
            ws.cell(row=notes_row, column=1).value = notes_text.strip()
            ws.cell(row=notes_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            
            lines_needed = max(5, len(notes_text) // 40 + len(selected_notes))
            ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row + lines_needed, end_column=7)

            # 磁吸魔法：將公司章吸附到注意事項正下方
            stamp_row = notes_row + lines_needed + 1
            for img in ws._images:
                try:
                    if hasattr(img.anchor, '_from'):
                        original_row = img.anchor._from.row + 1
                        col_idx = img.anchor._from.col + 1
                        col_letter = get_column_letter(col_idx)
                    else:
                        match = re.match(r"([A-Z]+)(\d+)", str(img.anchor))
                        if match:
                            col_letter = match.group(1)
                            original_row = int(match.group(2))
                        else:
                            continue
                    
                    if original_row > 10:
                        img.anchor = f"{col_letter}{stamp_row}"
                except Exception as e:
                    pass

    # ========================================================
    # ✨ 終極暴力破解：存檔前強制掃描所有圖片並寫死尺寸
    # 您可以修改這裡的數字，直到匯出的比例看起來最完美為止
    # (數值代表像素，通常 100~200 左右是合理範圍)
    STAMP_WIDTH = 170   # 👈 在這裡修改寬度
    STAMP_HEIGHT = 145  # 👈 在這裡修改高度
    # ========================================================

    for img in ws._images:
        try:
            # 取得圖片目前的列數
            if hasattr(img.anchor, '_from'):
                row_idx = img.anchor._from.row + 1
            else:
                match = re.match(r"([A-Z]+)(\d+)", str(img.anchor))
                if match:
                    row_idx = int(match.group(2))
                else:
                    row_idx = 0
            
            # 如果這張圖在第10列以下 (確保不影響未來可能加在表頭的Logo)
            if row_idx > 10:
                img.width = STAMP_WIDTH
                img.height = STAMP_HEIGHT
        except Exception:
            pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output