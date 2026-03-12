import streamlit as st
import pandas as pd
import json
import glob
import os
from copy import copy

# 設定網頁標題與寬度
st.set_page_config(page_title="Form Maker 估價單系統", layout="wide")

# === 🚀 核心優化：CSS 魔法強制移除「載入變灰」與「無法點擊」狀態 ===
st.markdown(
    """
    <style>
    /* 移除表格與按鈕在更新時的半透明遮罩 */
    [data-testid="stDataFrame"], [data-testid="stDataEditor"], .stButton > button {
        opacity: 1 !important;
        pointer-events: auto !important;
        transition: none !important;
    }
    /* 隱藏右上角礙眼的 Running... 動畫，假裝完全沒有重新整理 */
    [data-testid="stStatusWidget"] {
        display: none !important;
    }
    </style>    
    """,
    unsafe_allow_html=True
)
# =================================================================

@st.cache_data
def load_data():
    items_by_category = {}
    all_items_flat = {}
    filepaths = glob.glob("Item_Price/*_item.json")
    for path in filepaths:
        category = os.path.basename(path).replace("_item.json", "")
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            items_by_category[category] = data
            for item_name, details in data.items():
                all_items_flat[item_name] = {"price": details["price"], "unit": details["unit"], "category": category}
    return items_by_category, all_items_flat

items_by_category, all_items_flat = load_data()

if 'cart' not in st.session_state:
    st.session_state.cart = []

# 🚀 核心優化：改為 Callback 函數，這樣點擊按鈕的瞬間就會改變資料，不會有延遲感
# 🚀 核心優化：修改 Callback 函數，支援傳入自訂數量 (qty)
def add_to_cart(name, price, unit, qty=1):
    st.session_state.cart.append({
        "品名": name,
        "數量": qty,
        "單位": unit,
        "單價": int(price),
        "金額": int(price) * qty # 金額直接幫他算好
    })

def clear_cart():
    st.session_state.cart = []

def add_category_to_cart(category_name):
    items_in_cat = items_by_category[category_name]
    for name, details in items_in_cat.items():
        st.session_state.cart.append({
            "品名": name,
            "數量": 1,
            "單位": details["unit"],
            "單價": int(details["price"]),
            "金額": int(details["price"])
        })


# ================= 畫面 UI 佈局 =================

st.title("📝 Form Maker 智慧估價單")

col_left, col_right = st.columns([1, 1])

with col_left:
    st.header("1. 新增項目")
    
    # --- 方案 A: 關鍵字快速搜尋 ---
    st.subheader("🔍 快速搜尋")
    search_options = list(all_items_flat.keys())
    selected_item_search = st.selectbox("輸入關鍵字搜尋品名：", [""] + search_options, key="search_box")
    
    if selected_item_search != "":
        item_info = all_items_flat[selected_item_search]
        
        # 💡 新增：加入前微調區塊 (搜尋區)
        st.caption(f"📍 分類來源：{item_info['category']}")
        s_col1, s_col2, s_col3 = st.columns([3, 1, 2])
        # 帶入預設值，允許修改
        mod_name_s = s_col1.text_input("品名", value=selected_item_search, key="mod_name_s")
        mod_qty_s = s_col2.number_input("數量", value=1, min_value=1, step=1, key="mod_qty_s")
        mod_price_s = s_col3.number_input(f"單價 (/{item_info['unit']})", value=int(item_info['price']), min_value=0, step=100, key="mod_price_s")
        
        # 將修改後的值傳給 add_to_cart
        st.button("➕ 從搜尋加入", key="btn_search_add", type="primary", use_container_width=True, 
                  on_click=add_to_cart, args=(mod_name_s, mod_price_s, item_info['unit'], mod_qty_s))

    st.divider()

    # --- 方案 B: 分類聯動下拉選單 ---
    st.subheader("📁 依分類選擇")
    
    category_list = list(items_by_category.keys())
    selected_category = st.selectbox("步驟 1：選擇分類", ["請選擇分類..."] + category_list)
    
    if selected_category != "請選擇分類...":
        items_in_cat = items_by_category[selected_category]
        item_names = list(items_in_cat.keys())
        
        # 一鍵加入整個分類的按鈕 (維持原樣)
        st.button(f"⚡ 一鍵加入【{selected_category}】全品項", key="btn_add_all_cat", type="secondary", use_container_width=True, 
                  on_click=add_category_to_cart, args=(selected_category,))
        
        st.markdown("<p style='text-align: center; color: gray; margin-top: 10px; margin-bottom: 10px;'>— 或挑選單項 —</p>", unsafe_allow_html=True)
        
        selected_item_cat = st.selectbox("步驟 2：選擇單一品項", ["請選擇品項..."] + item_names)
        
        if selected_item_cat != "請選擇品項...":
            details = items_in_cat[selected_item_cat]
            
            # 💡 新增：加入前微調區塊 (分類區)
            c_col1, c_col2, c_col3 = st.columns([3, 1, 2])
            # 帶入預設值，允許修改
            mod_name_c = c_col1.text_input("品名", value=selected_item_cat, key="mod_name_c")
            mod_qty_c = c_col2.number_input("數量", value=1, min_value=1, step=1, key="mod_qty_c")
            mod_price_c = c_col3.number_input(f"單價 (/{details['unit']})", value=int(details['price']), min_value=0, step=100, key="mod_price_c")
            
            # 將修改後的值傳給 add_to_cart
            st.button("➕ 從分類加入單項", key="btn_cat_add", type="primary", use_container_width=True, 
                      on_click=add_to_cart, args=(mod_name_c, mod_price_c, details['unit'], mod_qty_c))

with col_right:
    st.header("2. 目前估價單預覽")
    
    if len(st.session_state.cart) > 0:
        df = pd.DataFrame(st.session_state.cart)
        
        st.write("💡 提示：雙擊數字即可出現上下微調按鈕，或點選後直接使用鍵盤「上下方向鍵」修改。")
        
        # 開放品名修改，並設定數值級距
        edited_df = st.data_editor(
            df,
            column_config={
                "品名": st.column_config.TextColumn(disabled=False), # 改為 False：開放鍵盤修改
                "單價": st.column_config.NumberColumn(disabled=False, step=100, min_value=0),
                "單位": st.column_config.TextColumn(disabled=False), # 單位也順便開放讓您可以改
                "數量": st.column_config.NumberColumn(disabled=False, step=1, min_value=1),
                "金額": st.column_config.NumberColumn(disabled=True) # 金額交給程式算，鎖定不可手動改
            },
            hide_index=True,
            use_container_width=True,
            num_rows="dynamic",
            key="cart_editor"
        )
        
        # --- 處理連動更新與金額計算邏輯 ---
        new_cart = edited_df.to_dict('records')
        needs_rerun = False
        total_price = 0
        
        for i, item in enumerate(new_cart):
            # 防呆：處理因為 dynamic 允許客戶手動新增空白列時，所產生的空值 (None)
            if pd.isna(item.get('品名')): item['品名'] = ""
            if pd.isna(item.get('數量')): item['數量'] = 1
            if pd.isna(item.get('單價')): item['單價'] = 0
            if pd.isna(item.get('單位')): item['單位'] = "式"
            
            # 計算正確的金額
            calc_total = int(item['數量'] * item['單價'])
            
            # 檢查：如果剛算出來的金額跟表格裡的不一樣，代表使用者改了數量或單價
            if calc_total != item.get('金額'):
                item['金額'] = calc_total
                needs_rerun = True
            
            # 檢查：是否有修改品名或單位
            if i < len(st.session_state.cart):
                old_item = st.session_state.cart[i]
                if item['品名'] != old_item['品名'] or item['單位'] != old_item['單位']:
                    needs_rerun = True
                    
            total_price += calc_total

        # 檢查：是否有刪除或新增列
        if len(new_cart) != len(st.session_state.cart):
            needs_rerun = True

        # 如果偵測到任何異動，就把新資料存入 session_state 並瞬間重整畫面以更新「金額」
        if needs_rerun:
            st.session_state.cart = new_cart
            st.rerun() 
        else:
            st.session_state.cart = new_cart
        
        st.divider()
        st.markdown(f"### 總計金額： **${total_price:,}**")
        
        st.button("🗑️ 清空估價單", type="primary", on_click=clear_cart)
            
    else:
        st.info("👈 請從左側選擇分類加入項目，或在下方直接新增")

        # ================= 匯出估價單區塊 =================
import io
import openpyxl
import datetime

st.divider()
st.header("3. 匯出估價單")

# 如果購物車內有東西，才顯示匯出選項
if len(st.session_state.cart) > 0:
    # 使用三欄式排版讓 UI 看起來整齊
    col_ex1, col_ex2, col_ex3 = st.columns(3)
    
    with col_ex1:
        client_name = st.text_input("客戶名稱 (TO)：", placeholder="例如：王大明 先生")
    
    with col_ex2:
        export_date = st.date_input("報價日期：", datetime.date.today())
    
    with col_ex3:
        tax_type = st.radio("營業稅計算：", ["已含稅", "未稅 (+5% 營業稅)"], horizontal=True)

    # 處理稅金與總額計算
    # total_price 是從上面右邊的預覽表格中算出來的原始總計
    subtotal = total_price
    
    if tax_type == "未稅 (+5% 營業稅)":
        tax = int(subtotal * 0.05)
        grand_total = subtotal + tax
    else:
        # 已含稅：將目前的總計視為「最後總價」，反推小計與稅金
        grand_total = subtotal
        subtotal = int(grand_total / 1.05)
        tax = grand_total - subtotal
        
    st.info(f"📊 **估價單試算** ➡️ 小計：${subtotal:,} | 營業稅：${tax:,} | 總計：${grand_total:,}")

    from copy import copy
    import io
    import openpyxl

    # --- 產生 Excel 檔案的邏輯 (完美終極版：精準不誤刪) ---
    def generate_excel():
        from openpyxl.styles import Alignment
        
        wb = openpyxl.load_workbook("blank_form.xlsx")
        ws = wb.active
        
        # 1. 填寫 TO 與日期 (固定在第 6 列)
        ws.cell(row=6, column=2).value = client_name
        
        minguo_year = export_date.year - 1911
        date_str = f"{minguo_year}/{export_date.month}/{export_date.day}"
        ws.cell(row=6, column=6).value = date_str
        
        # 2. 填寫購物車內的項目
        current_row = 8
        for idx, item in enumerate(st.session_state.cart):
            # 防呆：避免寫入到小計區塊
            if "小計" in str(ws.cell(row=current_row, column=5).value).strip():
                st.warning("項目數量超過表單預留空間，部分項目可能未匯出。")
                break
                
            ws.cell(row=current_row, column=1).value = idx + 1               # A欄：項次
            ws.cell(row=current_row, column=2).value = item.get("品名", "")  # B欄：品名
            ws.cell(row=current_row, column=3).value = item.get("數量", 0)   # C欄：數量
            ws.cell(row=current_row, column=4).value = item.get("單位", "")  # D欄：單位
            ws.cell(row=current_row, column=5).value = item.get("單價", 0)   # E欄：單價
            ws.cell(row=current_row, column=6).value = item.get("金額", 0)   # F欄：金額
            current_row += 1

        # 3. 尋找底部的「小計」(位在 E 欄)
        subtotal_row = -1
        for r in range(current_row, current_row + 100):
            cell_val = str(ws.cell(row=r, column=5).value).strip()
            
            if "小計" in cell_val:
                subtotal_row = r
                
                # 直接寫入純數值到 F 欄，覆寫掉 Excel 裡的公式
                ws.cell(row=r, column=6).value = subtotal       # 小計金額
                ws.cell(row=r+1, column=6).value = tax          # 營業稅金額
                ws.cell(row=r+2, column=6).value = grand_total  # 總和金額
                break

        # 4. 完美接合與動態對齊格式
        if subtotal_row != -1 and subtotal_row > current_row:
            # 刪除中間多餘的空白列
            rows_to_delete = subtotal_row - current_row
            ws.delete_rows(current_row, amount=rows_to_delete)
            
            # 此時，小計已經往上移動到了 current_row 的位置
            new_subtotal_row = current_row
            
            # (1) 讓「小計、營業稅、總和」完美靠右貼齊單價
            right_align = Alignment(horizontal='right')
            for i in range(3):
                ws.cell(row=new_subtotal_row + i, column=5).alignment = right_align
            
            # (2) 尋找注意事項並合併
            # 💡 關鍵修正：總和在 new_subtotal_row + 2，所以注意事項必須從 +3 開始找，才不會把總和吃掉！
            notes_row = -1
            for r in range(new_subtotal_row + 3, new_subtotal_row + 15):
                if str(ws.cell(row=r, column=1).value).strip() != "":
                    notes_row = r
                    break
            
            if notes_row != -1:
                # 靠上對齊並允許自動換行
                ws.cell(row=notes_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
                # 將注意事項從 A 欄橫向合併到 G 欄，並往下合併 5 行確保有足夠空間顯示文字
                ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row+5, end_column=7)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    excel_data = generate_excel() # 呼叫剛剛寫好的函數
    file_name = f"{client_name if client_name else '未命名'}_估價單.xlsx"
    
    st.download_button(
        label="📥 匯出並下載 Excel",
        data=excel_data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )