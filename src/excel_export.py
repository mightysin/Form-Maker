import io
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
import re
from copy import copy
import streamlit as st

def copy_worksheet_format_and_data(source_ws, target_ws):
    for mcr in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(mcr))
        
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width
        
    for row_idx, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = row_dim.height
            
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
                
    for img in source_ws._images:
        new_img = copy(img)
        new_img.width = img.width
        new_img.height = img.height
        target_ws.add_image(new_img, img.anchor)

def generate_excel(client_name, export_date, subtotal, tax, grand_total, cart, selected_notes, uploaded_file=None):
    template_wb = openpyxl.load_workbook("blank_form.xlsx")
    source_ws = template_wb.active
    
    if uploaded_file is not None:
        wb = openpyxl.load_workbook(uploaded_file)
        sheet_title = f"{export_date.month:02d}{export_date.day:02d}_報價"
        base_title = sheet_title
        counter = 1
        while sheet_title in wb.sheetnames:
            sheet_title = f"{base_title}_{counter}"
            counter += 1
        ws = wb.create_sheet(title=sheet_title)
        copy_worksheet_format_and_data(source_ws, ws)
        wb.active = ws  
    else:
        wb = template_wb
        ws = wb.active
        if client_name:
            ws.title = str(client_name)[:31] 

    ws.cell(row=6, column=2).value = client_name
    
    minguo_year = export_date.year - 1911
    date_str = f"{minguo_year}/{export_date.month}/{export_date.day}"
    ws.cell(row=6, column=6).value = date_str
    
    current_row = 8
    for idx, item in enumerate(cart):
        if "小計" in str(ws.cell(row=current_row, column=5).value).strip():
            st.warning("項目數量超過表單預留空間，部分項目可能未匯出。")
            break
            
        ws.cell(row=current_row, column=1).value = idx + 1               
        ws.cell(row=current_row, column=2).value = item.get("品名", "")  
        
        try: qty = int(float(item.get("數量", 0) or 0))
        except Exception: qty = 0
        
        try: price = int(float(item.get("單價", 0) or 0))
        except Exception: price = 0
        
        ws.cell(row=current_row, column=3).value = qty   
        ws.cell(row=current_row, column=4).value = item.get("單位", "")  
        ws.cell(row=current_row, column=5).value = price   
        
        # ✨ 重點修改 1：不寫死金額，而是寫入 Excel 公式 (C欄 乘 E欄)
        ws.cell(row=current_row, column=6).value = f"=C{current_row}*E{current_row}"   
        current_row += 1

    subtotal_row = -1
    for r in range(current_row, current_row + 100):
        # 加強搜尋：同時找 D 欄跟 E 欄，避免抓不到小計位置
        if "小計" in str(ws.cell(row=r, column=4).value).strip() or "小計" in str(ws.cell(row=r, column=5).value).strip():
            subtotal_row = r
            break

    if subtotal_row != -1 and subtotal_row > current_row:
        # 1. 解除合併儲存格
        ranges_to_unmerge = []
        for m_range in ws.merged_cells.ranges:
            if not (m_range.max_row < current_row or m_range.min_row >= subtotal_row):
                ranges_to_unmerge.append(m_range)
        for m_range in ranges_to_unmerge:
            ws.unmerge_cells(str(m_range))
            
        # 2. 刪除多餘空白列
        rows_to_delete = subtotal_row - current_row
        ws.delete_rows(current_row, amount=rows_to_delete)
        new_subtotal_row = current_row
        
        # 3. 先畫框線與背景色 (避免覆蓋到後面的文字)
        right_align = Alignment(horizontal='right', vertical='center')
        thin_side = Side(border_style="thin", color="000000")
        no_side = Side(border_style=None)
        
        border_left = Border(top=thin_side, bottom=thin_side, left=thin_side, right=no_side)
        border_middle = Border(top=thin_side, bottom=thin_side, left=no_side, right=no_side)
        border_right = Border(top=thin_side, bottom=thin_side, left=no_side, right=thin_side)
        border_full = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        white_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
        bold_font = Font(name="新細明體", size=14, bold=True) # 設定字體加粗
        
        for i in range(3):
            row_idx = new_subtotal_row + i
            for col_idx in range(1, 7):
                target_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell.fill = white_fill
                if col_idx == 1: target_cell.border = border_left
                elif col_idx in [2, 3, 4]: target_cell.border = border_middle
                elif col_idx == 5: target_cell.border = border_right
                elif col_idx == 6: target_cell.border = border_full

        # 4. ✨ 暴力強制寫入文字與公式 (畫完格子後再寫，保證不會消失)
        last_item_row = current_row - 1
        sum_range = f"F8:F{last_item_row}" if last_item_row >= 8 else "0"
        raw_sum = sum([(int(float(i.get("數量", 0) or 0)) * int(float(i.get("單價", 0) or 0))) for i in cart])

        # 寫入文字標籤並設定對齊
        ws.cell(row=new_subtotal_row, column=5).value = "小計"
        ws.cell(row=new_subtotal_row+2, column=5).value = "總計"
        
        for i in range(3):
            ws.cell(row=new_subtotal_row + i, column=5).alignment = right_align
            ws.cell(row=new_subtotal_row + i, column=5).font = bold_font
            ws.cell(row=new_subtotal_row + i, column=6).font = bold_font

        if raw_sum == grand_total and tax > 0:
            ws.cell(row=new_subtotal_row+1, column=5).value = "營業稅(內含)"
            ws.cell(row=new_subtotal_row+2, column=6).value = f"=SUM({sum_range})"
            ws.cell(row=new_subtotal_row, column=6).value = f"=ROUND(F{new_subtotal_row+2}/1.05, 0)"
            ws.cell(row=new_subtotal_row+1, column=6).value = f"=F{new_subtotal_row+2}-F{new_subtotal_row}"
        else:
            ws.cell(row=new_subtotal_row+1, column=5).value = "營業稅(5%)"
            ws.cell(row=new_subtotal_row, column=6).value = f"=SUM({sum_range})"
            ws.cell(row=new_subtotal_row+1, column=6).value = f"=ROUND(F{new_subtotal_row}*0.05, 0)"
            ws.cell(row=new_subtotal_row+2, column=6).value = f"=F{new_subtotal_row}+F{new_subtotal_row+1}"
        
        notes_row = -1
        for r in range(new_subtotal_row + 3, new_subtotal_row + 15):
            if str(ws.cell(row=r, column=1).value).strip() != "":
                notes_row = r
                break
        
        if notes_row != -1:
            notes_text = ""
            for i, note in enumerate(selected_notes):
                notes_text += f"{i+1}. {note}\n"
            
            ws.cell(row=notes_row, column=1).value = notes_text.strip()
            ws.cell(row=notes_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            
            lines_needed = max(5, len(notes_text) // 40 + len(selected_notes))
            ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row + lines_needed, end_column=7)

            stamp_row = notes_row + lines_needed + 1
            for img in ws._images:
                try:
                    orig_w = img.width
                    orig_h = img.height
                    
                    if hasattr(img.anchor, '_from'):
                        original_row = img.anchor._from.row + 1
                        col_idx = img.anchor._from.col + 1
                        col_letter = get_column_letter(col_idx)
                    else:
                        match = re.match(r"([A-Z]+)(\d+)", str(img.anchor))
                        if match:
                            col_letter = match.group(1)
                            original_row = int(match.group(2))
                        else:
                            continue
                    
                    if original_row > 10:
                        img.anchor = f"{col_letter}{stamp_row}"
                        img.width = orig_w
                        img.height = orig_h
                except Exception:
                    pass
                    
    STAMP_WIDTH = 180
    STAMP_HEIGHT = 180
    for img in ws._images:
        try:
            if hasattr(img.anchor, '_from'):
                row_idx = img.anchor._from.row + 1
            else:
                match = re.match(r"([A-Z]+)(\d+)", str(img.anchor))
                if match:
                    row_idx = int(match.group(2))
                else:
                    row_idx = 0
            
            if row_idx > 10:
                img.width = STAMP_WIDTH
                img.height = STAMP_HEIGHT
        except Exception:
            pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output